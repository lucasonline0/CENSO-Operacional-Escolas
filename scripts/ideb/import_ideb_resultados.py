# -*- coding: utf-8 -*-
"""
IDEB-03A — Importador (dry-run) da base IDEB 2023 para a tabela ideb_resultados.

Este script lê a planilha oficial do IDEB 2023 (insumo local, NÃO versionado),
valida colunas, normaliza campos, classifica status_ideb / detalhe_status_ideb,
prepara o vínculo com schools por codigo_inep e gera um relatório de dry-run.

Modo padrão: --dry-run (seguro, NÃO escreve no banco).
A carga real (--apply) é deliberadamente travada por confirmação extra e é
escopo de um incremento POSTERIOR (IDEB-03B). Não use --apply nesta etapa.

Regras metodológicas (docs/dashboard/perfil-alunos-resultados-ideb-2023.md e
infra/migrations/0017_create_ideb_resultados.sql):
  * codigo_inep é a CHAVE de integração; preservado como TEXTO (zeros à esquerda,
    sem sufixo ".0" herdado do Excel).
  * '-', '', 'ND' => NULL numérico, NUNCA 0. Ausência é cobertura/elegibilidade.
  * status_ideb guarda-chuva: com_ideb | sem_ideb_divulgado.
  * detalhe_status_ideb: NULL (quando com_ideb) | sem_resultado | nd_proficiencia | outro.
  * percentual_avaliado > 100 é preservado (apenas alerta de qualidade).
  * agregações por DRE/município NÃO são IDEB oficial agregado do INEP.

Exemplos de uso:

  # Dry-run (padrão; não escreve nada):
  python scripts/ideb/import_ideb_resultados.py \
      --source _local/ideb/fontes/ideb_2023_iniciais_finais_medio.xlsx \
      --ano 2023 --dry-run

  # Carga real (FUTURO — IDEB-03B; exige --confirm-apply):
  python scripts/ideb/import_ideb_resultados.py \
      --source _local/ideb/fontes/ideb_2023_iniciais_finais_medio.xlsx \
      --ano 2023 --apply --confirm-apply \
      --batch-id ideb_2023_20260614_120000
"""

import argparse
import json
import os
import re
import sys
import unicodedata
from datetime import datetime

# openpyxl é a única dependência obrigatória para o dry-run.
try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - orientação amigável
    print(
        "[ERRO] openpyxl não está instalado.\n"
        "       Instale as dependências com:\n"
        "         pip install -r scripts/ideb/requirements.txt",
        file=sys.stderr,
    )
    sys.exit(2)

# psycopg é OPCIONAL no dry-run (só usado para simular o match com schools).
try:
    import psycopg  # type: ignore

    _HAS_PSYCOPG = True
except ImportError:
    psycopg = None  # type: ignore
    _HAS_PSYCOPG = False


# ---------------------------------------------------------------------------
# Constantes da fonte
# ---------------------------------------------------------------------------
ABA = "IDEB 2023"
FONTE_ARQUIVO_PADRAO = "ideb_2023_iniciais_finais_medio.xlsx"
FONTE_INEP_URL = "https://download.inep.gov.br/ideb/nota_informativa_ideb_2023.pdf"

RELATORIO_MD = "_local/ideb/relatorios/import_dry_run_ideb_2023.md"
RELATORIO_JSON = "_local/ideb/relatorios/import_dry_run_ideb_2023.json"

# Marcadores textuais que indicam ausência de valor numérico (nunca viram zero).
MARCADORES_AUSENCIA = {"-", "", "nd", "n/d", "na", "n/a", "nan", "none"}

# Colunas esperadas: alias lógico -> rótulo canônico na planilha.
COLUNAS = {
    "ano": "ANO",
    "ensino": "Ensino",
    "inep": "INEP",
    "nome_escola": "NOME DA ESCOLA",
    "total_avaliado": "Total avaliado",
    "percentual_avaliado": "Percentual avaliado",
    "proficiencia_portugues": "Proficiência Português",
    "proficiencia_matematica": "Proficiência Matemática",
    "fluxo_indicador_rendimento": "Fluxo - Indicador de rendimento",
    "ideb": "IDEB 2023",
}

# Campos numéricos (na ordem da planilha) que descrevem "todos os indicadores"
# de Total avaliado até IDEB.
CAMPOS_INDICADORES = [
    "total_avaliado",
    "percentual_avaliado",
    "proficiencia_portugues",
    "proficiencia_matematica",
    "fluxo_indicador_rendimento",
    "ideb",
]

ETAPAS_ORDEM = ["anos_iniciais", "anos_finais", "ensino_medio"]
ETAPA_LABEL = {
    "anos_iniciais": "Anos iniciais",
    "anos_finais": "Anos finais",
    "ensino_medio": "Ensino médio",
    "desconhecida": "Desconhecida",
}


# ---------------------------------------------------------------------------
# Normalização e parsing (idênticos às regras do IDEB-01 / _common.py)
# ---------------------------------------------------------------------------
def norm(texto):
    """Normaliza texto para comparação tolerante (sem acento, minúsculo, sem espaço extra)."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return " ".join(t.lower().split())


def texto_limpo(valor):
    if valor is None:
        return ""
    if isinstance(valor, float):
        # openpyxl pode devolver float; reconstrói texto sem ".0" espúrio.
        if valor != valor:  # NaN
            return ""
        if valor.is_integer():
            return str(int(valor))
        return str(valor)
    return str(valor).strip()


def is_ausente(valor):
    """True se o valor representa ausência (-, ND, vazio, NaN). Nunca vira zero."""
    return norm(texto_limpo(valor)) in MARCADORES_AUSENCIA


def is_nd(valor):
    """True se o valor é explicitamente 'ND' / 'N/D' (subtipo de ausência)."""
    return norm(texto_limpo(valor)) in {"nd", "n/d"}


def parse_num(valor):
    """Converte para float quando numérico; caso contrário (incl. '-' e 'ND') -> None.

    Aceita vírgula decimal. NUNCA retorna 0 para ausência.
    """
    if isinstance(valor, (int, float)) and not (isinstance(valor, float) and valor != valor):
        return float(valor)
    t = texto_limpo(valor)
    if is_ausente(t):
        return None
    t = t.replace(" ", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return None


def normalizar_etapa(ensino):
    """Normaliza o campo Ensino para chave canônica (tolerante a acento/caixa/espaço)."""
    n = norm(ensino)
    if "inicia" in n:
        return "anos_iniciais"
    if "fina" in n:  # cobre "anos finais" e "anos final"
        return "anos_finais"
    if "medio" in n:
        return "ensino_medio"
    return "desconhecida"


def normalizar_inep(valor):
    """Preserva INEP como TEXTO, removendo sufixo '.0' do Excel e espaços.

    Não converte para inteiro: zeros à esquerda devem ser preservados.
    """
    t = texto_limpo(valor)
    if t.endswith(".0") and t[:-2].isdigit():
        t = t[:-2]
    return t


def classificar_status(linha):
    """Retorna (status_ideb, detalhe_status_ideb) conforme o documento metodológico.

    status_ideb:
        com_ideb            : IDEB 2023 numérico.
        sem_ideb_divulgado  : IDEB ausente / não numérico.

    detalhe_status_ideb (somente quando sem_ideb_divulgado):
        sem_resultado   : todos os indicadores (Total avaliado..IDEB) ausentes/'-'.
        nd_proficiencia : IDEB ausente e alguma proficiência == 'ND'.
        outro           : demais casos sem IDEB.
    Quando com_ideb, detalhe_status_ideb é None.
    """
    if linha["ideb"] is not None:
        return "com_ideb", None

    todos_ausentes = all(is_ausente(linha["_raw"][c]) for c in CAMPOS_INDICADORES)
    if todos_ausentes:
        return "sem_ideb_divulgado", "sem_resultado"

    prof_nd = is_nd(linha["_raw"]["proficiencia_portugues"]) or is_nd(
        linha["_raw"]["proficiencia_matematica"]
    )
    if prof_nd:
        return "sem_ideb_divulgado", "nd_proficiencia"

    return "sem_ideb_divulgado", "outro"


# ---------------------------------------------------------------------------
# Leitura da planilha (openpyxl, preservando texto)
# ---------------------------------------------------------------------------
def carregar_planilha(path):
    """Lê a aba 'IDEB 2023' e retorna (header_map, registros_brutos).

    header_map: alias_lógico -> índice de coluna (0-based).
    registros_brutos: lista de dicts alias_lógico -> valor cru da célula.
    Levanta KeyError se faltar coluna esperada.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        raise KeyError(
            f"Aba '{ABA}' não encontrada na planilha. Abas presentes: {wb.sheetnames}"
        )
    ws = wb[ABA]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise KeyError("Planilha vazia: nenhuma linha de cabeçalho encontrada.")

    norm_para_idx = {}
    for idx, nome in enumerate(header):
        if nome is None:
            continue
        norm_para_idx[norm(nome)] = idx

    header_map = {}
    faltando = []
    for alias, esperada in COLUNAS.items():
        idx = norm_para_idx.get(norm(esperada))
        if idx is None:
            faltando.append(esperada)
        else:
            header_map[alias] = idx
    if faltando:
        raise KeyError(f"Colunas esperadas ausentes na planilha (aba '{ABA}'): {faltando}")

    registros = []
    for row in rows:
        if row is None:
            continue
        # Pula linhas totalmente vazias.
        if all(c is None or texto_limpo(c) == "" for c in row):
            continue
        reg = {}
        for alias, idx in header_map.items():
            reg[alias] = row[idx] if idx < len(row) else None
        registros.append(reg)

    wb.close()
    return header_map, registros


def normalizar_registro(reg, ano_param):
    """Transforma um registro bruto em registro normalizado pronto para análise/carga."""
    linha = {
        "ano": ano_param,
        "ano_origem": texto_limpo(reg["ano"]),
        "etapa": normalizar_etapa(reg["ensino"]),
        "codigo_inep": normalizar_inep(reg["inep"]),
        "nome_escola_origem": texto_limpo(reg["nome_escola"]),
        "total_avaliado": parse_num(reg["total_avaliado"]),
        "percentual_avaliado": parse_num(reg["percentual_avaliado"]),
        "proficiencia_portugues": parse_num(reg["proficiencia_portugues"]),
        "proficiencia_matematica": parse_num(reg["proficiencia_matematica"]),
        "fluxo_indicador_rendimento": parse_num(reg["fluxo_indicador_rendimento"]),
        "ideb": parse_num(reg["ideb"]),
        "_raw": reg,
    }
    status_ideb, detalhe = classificar_status(linha)
    linha["status_ideb"] = status_ideb
    linha["detalhe_status_ideb"] = detalhe
    return linha


# ---------------------------------------------------------------------------
# Conexão com banco (somente leitura no dry-run; mascarando segredos)
# ---------------------------------------------------------------------------
def carregar_dotenv_best_effort():
    """Carrega variáveis de um .env, se existir, sem depender de libs externas.

    Procura nos mesmos lugares que o backend Go (raiz, infra/). Não sobrescreve
    variáveis já definidas no ambiente.
    """
    candidatos = [".env", os.path.join("infra", ".env"), os.path.join("..", ".env")]
    for caminho in candidatos:
        if not os.path.isfile(caminho):
            continue
        try:
            with open(caminho, "r", encoding="utf-8") as f:
                for raw in f:
                    linha = raw.strip()
                    if not linha or linha.startswith("#") or "=" not in linha:
                        continue
                    chave, _, valor = linha.partition("=")
                    chave = chave.strip()
                    valor = valor.strip().strip('"').strip("'")
                    if chave and chave not in os.environ:
                        os.environ[chave] = valor
        except OSError:
            continue


def resolver_dsn(dsn_flag):
    """Resolve o DSN: --dsn > DATABASE_URL > DB_DSN > componentes DB_HOST/...

    Espelha a resolução do backend (api/cmd/import-prodep/main.go). Retorna ""
    quando nenhuma fonte está disponível.
    """
    carregar_dotenv_best_effort()
    if dsn_flag:
        return dsn_flag
    if os.environ.get("DATABASE_URL"):
        return os.environ["DATABASE_URL"]
    if os.environ.get("DB_DSN"):
        return os.environ["DB_DSN"]
    host = os.environ.get("DB_HOST")
    if host:
        sslmode = os.environ.get("DB_SSLMODE") or "disable"
        return (
            f"host={host} port={os.environ.get('DB_PORT', '')} "
            f"user={os.environ.get('DB_USER', '')} password={os.environ.get('DB_PASSWORD', '')} "
            f"dbname={os.environ.get('DB_NAME', '')} sslmode={sslmode} connect_timeout=5"
        )
    return ""


def mascarar_dsn(dsn):
    """Mascara segredos do DSN para log. Nunca imprime senha, host completo ou token."""
    if not dsn:
        return "(sem DSN)"
    # Formato URL: postgresql://user:pass@host:port/db
    m = re.match(r"^(?P<scheme>\w+)://(?P<userinfo>[^@/]*)@?(?P<rest>.*)$", dsn)
    if m and "://" in dsn:
        scheme = m.group("scheme")
        host_part = m.group("rest").split("/")[0].split("@")[-1]
        host = host_part.split(":")[0] if host_part else "?"
        host_mask = host[:3] + "***" if len(host) > 3 else "***"
        return f"{scheme}://***:***@{host_mask}/*** (mascarado)"
    # Formato key=value: mascara password e revela só prefixo do host.
    partes = []
    for token in dsn.split():
        if "=" not in token:
            continue
        chave, _, valor = token.partition("=")
        if chave.lower() in {"password", "user"}:
            partes.append(f"{chave}=***")
        elif chave.lower() == "host":
            partes.append(f"host={valor[:3]}***" if len(valor) > 3 else "host=***")
        else:
            partes.append(f"{chave}={valor}")
    return " ".join(partes) + " (mascarado)"


def consultar_schools_por_inep(dsn, ineps):
    """Consulta schools (somente SELECT) e retorna dict codigo_inep -> {id, nome_escola}.

    Usa lotes parametrizados (= ANY($1)). Nunca escreve nada.
    """
    resultado = {}
    if not dsn or not _HAS_PSYCOPG or not ineps:
        return resultado
    ineps_lista = sorted(ineps)
    with psycopg.connect(dsn, connect_timeout=10) as conn:  # type: ignore[union-attr]
        with conn.cursor() as cur:
            cur.execute(
                "SELECT codigo_inep, id, nome_escola FROM schools "
                "WHERE codigo_inep = ANY(%s)",
                (ineps_lista,),
            )
            for codigo_inep, school_id, nome_escola in cur.fetchall():
                if codigo_inep is None:
                    continue
                resultado[normalizar_inep(codigo_inep)] = {
                    "id": school_id,
                    "nome_escola": nome_escola or "",
                }
    return resultado


# ---------------------------------------------------------------------------
# Resolução de vínculo (match com schools)
# ---------------------------------------------------------------------------
def resolver_vinculo(linhas, schools_por_inep, houve_conexao):
    """Preenche school_id, status_vinculo e coleta alertas de divergência de nome.

    Sem conexão: school_id=NULL, status_vinculo=pendente_validacao (match não executado).
    Com conexão:
        match  -> school_id preenchido, status_vinculo=match_inep;
        sem match -> school_id=NULL, status_vinculo=sem_match_inep.
    Divergência de nome com mesmo INEP é apenas alerta, nunca bloqueio.
    """
    alertas_nome = []
    for l in linhas:
        if not houve_conexao:
            l["school_id"] = None
            l["status_vinculo"] = "pendente_validacao"
            continue
        match = schools_por_inep.get(l["codigo_inep"])
        if match:
            l["school_id"] = match["id"]
            l["status_vinculo"] = "match_inep"
            nome_ideb = norm(l["nome_escola_origem"])
            nome_school = norm(match["nome_escola"])
            if nome_ideb and nome_school and nome_ideb != nome_school:
                alertas_nome.append(
                    {
                        "codigo_inep": l["codigo_inep"],
                        "nome_ideb": l["nome_escola_origem"],
                        "nome_schools": match["nome_escola"],
                    }
                )
        else:
            l["school_id"] = None
            l["status_vinculo"] = "sem_match_inep"
    return alertas_nome


# ---------------------------------------------------------------------------
# Estatísticas / relatório
# ---------------------------------------------------------------------------
def faixa_de(v):
    if v is None:
        return "Sem IDEB divulgado"
    if v < 3.0:
        return "Abaixo de 3,0"
    if v < 4.0:
        return "3,0 a 3,9"
    if v < 5.0:
        return "4,0 a 4,9"
    if v < 6.0:
        return "5,0 a 5,9"
    if v < 7.0:
        return "6,0 a 6,9"
    return "7,0+"


def fmt(v, casas=2):
    if v is None:
        return "—"
    return f"{v:.{casas}f}".replace(".", ",")


def calcular_estatisticas(linhas):
    """Calcula todos os números do relatório de dry-run."""
    total = len(linhas)
    ineps = sorted({l["codigo_inep"] for l in linhas if l["codigo_inep"]})

    com_ideb = [l for l in linhas if l["ideb"] is not None]
    sem_ideb = [l for l in linhas if l["ideb"] is None]

    por_etapa = {}
    for et in ETAPAS_ORDEM:
        sub = [l for l in linhas if l["etapa"] == et]
        por_etapa[et] = {
            "registros": len(sub),
            "ineps": len({l["codigo_inep"] for l in sub if l["codigo_inep"]}),
            "com_ideb": sum(1 for l in sub if l["ideb"] is not None),
            "sem_ideb": sum(1 for l in sub if l["ideb"] is None),
        }
    etapas_desconhecidas = sum(1 for l in linhas if l["etapa"] == "desconhecida")

    detalhe_count = {}
    for l in linhas:
        d = l["detalhe_status_ideb"]
        chave = d if d is not None else "(NULL / com_ideb)"
        detalhe_count[chave] = detalhe_count.get(chave, 0) + 1

    status_count = {}
    for l in linhas:
        status_count[l["status_ideb"]] = status_count.get(l["status_ideb"], 0) + 1

    reg_nd = sum(
        1
        for l in linhas
        if is_nd(l["_raw"]["proficiencia_portugues"]) or is_nd(l["_raw"]["proficiencia_matematica"])
    )
    perc_acima_100 = sum(
        1 for l in linhas if l["percentual_avaliado"] is not None and l["percentual_avaliado"] > 100
    )
    perc_abaixo_80 = sum(
        1 for l in linhas if l["percentual_avaliado"] is not None and l["percentual_avaliado"] < 80
    )

    # Duplicidades em ano + codigo_inep + etapa.
    chaves = {}
    for l in linhas:
        k = (l["ano"], l["codigo_inep"], l["etapa"])
        chaves[k] = chaves.get(k, 0) + 1
    duplicidades = {k: n for k, n in chaves.items() if n > 1}

    # Faixas por etapa.
    faixas = {}
    for et in ETAPAS_ORDEM:
        faixas[et] = {}
        for l in linhas:
            if l["etapa"] != et:
                continue
            f = faixa_de(l["ideb"])
            faixas[et][f] = faixas[et].get(f, 0) + 1

    return {
        "total_registros": total,
        "ineps_unicos": len(ineps),
        "com_ideb": len(com_ideb),
        "sem_ideb": len(sem_ideb),
        "por_etapa": por_etapa,
        "etapas_desconhecidas": etapas_desconhecidas,
        "detalhe_status_count": detalhe_count,
        "status_count": status_count,
        "registros_nd": reg_nd,
        "perc_acima_100": perc_acima_100,
        "perc_abaixo_80": perc_abaixo_80,
        "duplicidades": duplicidades,
        "faixas": faixas,
    }


def amostra_registros(linhas, n=10):
    """Retorna uma amostra de n registros normalizados (sem o campo bruto interno)."""
    saida = []
    for l in linhas[:n]:
        saida.append(
            {
                "ano": l["ano"],
                "codigo_inep": l["codigo_inep"],
                "etapa": l["etapa"],
                "nome_escola_origem": l["nome_escola_origem"],
                "total_avaliado": l["total_avaliado"],
                "percentual_avaliado": l["percentual_avaliado"],
                "proficiencia_portugues": l["proficiencia_portugues"],
                "proficiencia_matematica": l["proficiencia_matematica"],
                "fluxo_indicador_rendimento": l["fluxo_indicador_rendimento"],
                "ideb": l["ideb"],
                "status_ideb": l["status_ideb"],
                "detalhe_status_ideb": l["detalhe_status_ideb"],
                "school_id": l.get("school_id"),
                "status_vinculo": l.get("status_vinculo"),
            }
        )
    return saida


def gerar_relatorio_md(ctx):
    """Monta o relatório de dry-run em Markdown a partir do contexto consolidado."""
    s = ctx["stats"]
    L = []
    L.append("# Dry-run — Importador IDEB 2023 (IDEB-03A)")
    L.append("")
    L.append(f"> Gerado em {ctx['timestamp']}. Artefato **local** (não versionado).")
    L.append("> **Nenhuma escrita foi executada no banco.** Modo: dry-run.")
    L.append("")
    L.append("## Observações metodológicas")
    L.append("")
    L.append("- `-`, `ND` e ausência de IDEB **não** viram zero; viram `NULL`.")
    L.append("- IDEB ausente => `status_ideb = sem_ideb_divulgado` (cobertura/elegibilidade).")
    L.append("- `codigo_inep` preservado como **texto** (zeros à esquerda mantidos).")
    L.append("- `percentual_avaliado > 100` é preservado (apenas alerta de qualidade).")
    L.append("- Agregações por DRE/município **não** são IDEB oficial agregado do INEP.")
    L.append("")
    L.append("## 1. Execução")
    L.append("")
    L.append("| Item | Valor |")
    L.append("|---|---|")
    L.append(f"| Data/hora | {ctx['timestamp']} |")
    L.append(f"| Fonte | `{ctx['source']}` |")
    L.append(f"| Ano de referência | {ctx['ano']} |")
    L.append(f"| Modo | dry-run (sem escrita) |")
    L.append(f"| Conexão com banco | {'sim' if ctx['houve_conexao'] else 'não'} |")
    if ctx["dsn_mascarado"]:
        L.append(f"| DSN (mascarado) | `{ctx['dsn_mascarado']}` |")
    L.append("")
    L.append("## 2. Volumetria")
    L.append("")
    L.append("| Métrica | Valor |")
    L.append("|---|---:|")
    L.append(f"| Total de registros lidos | {s['total_registros']} |")
    L.append(f"| INEPs únicos | {s['ineps_unicos']} |")
    L.append(f"| Registros com IDEB | {s['com_ideb']} |")
    L.append(f"| Registros sem IDEB | {s['sem_ideb']} |")
    L.append(f"| Etapas desconhecidas (não mapeadas) | {s['etapas_desconhecidas']} |")
    L.append("")
    L.append("## 3. Registros por etapa")
    L.append("")
    L.append("| Etapa | Registros | INEPs | Com IDEB | Sem IDEB |")
    L.append("|---|---:|---:|---:|---:|")
    for et in ETAPAS_ORDEM:
        e = s["por_etapa"][et]
        L.append(
            f"| {ETAPA_LABEL[et]} | {e['registros']} | {e['ineps']} | "
            f"{e['com_ideb']} | {e['sem_ideb']} |"
        )
    L.append("")
    L.append("## 4. status_ideb e detalhe_status_ideb")
    L.append("")
    L.append("| status_ideb | Registros |")
    L.append("|---|---:|")
    for st in ["com_ideb", "sem_ideb_divulgado"]:
        L.append(f"| {st} | {s['status_count'].get(st, 0)} |")
    L.append("")
    L.append("| detalhe_status_ideb | Registros |")
    L.append("|---|---:|")
    for d in ["(NULL / com_ideb)", "sem_resultado", "nd_proficiencia", "outro"]:
        L.append(f"| {d} | {s['detalhe_status_count'].get(d, 0)} |")
    L.append("")
    L.append("## 5. Qualidade")
    L.append("")
    L.append("| Situação | Quantidade |")
    L.append("|---|---:|")
    L.append(f"| Registros com `ND` em proficiência | {s['registros_nd']} |")
    L.append(f"| Percentual avaliado acima de 100 | {s['perc_acima_100']} |")
    L.append(f"| Percentual avaliado abaixo de 80 | {s['perc_abaixo_80']} |")
    L.append(f"| Duplicidades em (ano + codigo_inep + etapa) | {len(s['duplicidades'])} |")
    L.append("")
    L.append("## 6. Distribuição por faixa de IDEB")
    L.append("")
    L.append("| Etapa | Faixa | Registros |")
    L.append("|---|---|---:|")
    ordem_faixas = [
        "Abaixo de 3,0",
        "3,0 a 3,9",
        "4,0 a 4,9",
        "5,0 a 5,9",
        "6,0 a 6,9",
        "7,0+",
        "Sem IDEB divulgado",
    ]
    for et in ETAPAS_ORDEM:
        for f in ordem_faixas:
            qtd = s["faixas"][et].get(f, 0)
            if qtd:
                L.append(f"| {ETAPA_LABEL[et]} | {f} | {qtd} |")
    L.append("")
    L.append("## 7. Vínculo com `schools`")
    L.append("")
    if ctx["houve_conexao"]:
        L.append("| Métrica | Valor |")
        L.append("|---|---:|")
        L.append(f"| Registros com match (match_inep) | {ctx['vinculo']['match_inep']} |")
        L.append(f"| Registros sem match (sem_match_inep) | {ctx['vinculo']['sem_match_inep']} |")
        L.append(f"| INEPs únicos com match | {ctx['vinculo']['ineps_match']} |")
        L.append(f"| INEPs únicos sem match | {ctx['vinculo']['ineps_sem_match']} |")
        L.append(f"| Alertas de divergência de nome (mesmo INEP) | {len(ctx['alertas_nome'])} |")
    else:
        L.append("Sem conexão com o banco nesta execução. O match real **não** foi executado.")
        L.append("")
        L.append(f"- Todos os {s['total_registros']} registros ficam com "
                 "`school_id = NULL` e `status_vinculo = pendente_validacao`.")
    L.append("")
    L.append("## 8. Amostra de registros normalizados (10)")
    L.append("")
    L.append("```json")
    L.append(json.dumps(ctx["amostra"], ensure_ascii=False, indent=2))
    L.append("```")
    L.append("")
    L.append("## 9. Alertas")
    L.append("")
    if ctx["alertas"]:
        for a in ctx["alertas"]:
            L.append(f"- {a}")
    else:
        L.append("- Nenhum alerta crítico.")
    if ctx["alertas_nome"]:
        L.append("")
        L.append("### Divergências de nome (mesmo INEP — apenas conferência)")
        L.append("")
        L.append("| INEP | Nome IDEB | Nome schools |")
        L.append("|---|---|---|")
        for a in ctx["alertas_nome"][:50]:
            L.append(f"| {a['codigo_inep']} | {a['nome_ideb']} | {a['nome_schools']} |")
        if len(ctx["alertas_nome"]) > 50:
            L.append(f"| ... | (+{len(ctx['alertas_nome']) - 50} divergências) | |")
    L.append("")
    L.append("## 10. Confirmação")
    L.append("")
    L.append("- ✅ Nenhuma escrita foi executada no banco (modo dry-run).")
    L.append("- ✅ INEP preservado como texto.")
    L.append("- ✅ `-` / `ND` / ausência convertidos para `NULL` (nunca 0).")
    L.append(f"- Próximo passo: **IDEB-03B** (carga controlada com `--apply --confirm-apply`).")
    L.append("")
    if s["duplicidades"]:
        L.append("## Anexo: duplicidades encontradas (ano + codigo_inep + etapa)")
        L.append("")
        L.append("| Ano | INEP | Etapa | Ocorrências |")
        L.append("|---|---|---|---:|")
        for (ano, inep, et), n in sorted(s["duplicidades"].items()):
            L.append(f"| {ano} | {inep} | {et} | {n} |")
        L.append("")
    return "\n".join(L) + "\n"


def gerar_relatorio_json(ctx):
    """Versão JSON do relatório (sem campos brutos internos)."""
    s = ctx["stats"]
    return {
        "timestamp": ctx["timestamp"],
        "source": ctx["source"],
        "ano": ctx["ano"],
        "modo": "dry-run",
        "houve_conexao": ctx["houve_conexao"],
        "escrita_executada": False,
        "volumetria": {
            "total_registros": s["total_registros"],
            "ineps_unicos": s["ineps_unicos"],
            "com_ideb": s["com_ideb"],
            "sem_ideb": s["sem_ideb"],
            "etapas_desconhecidas": s["etapas_desconhecidas"],
        },
        "por_etapa": s["por_etapa"],
        "status_count": s["status_count"],
        "detalhe_status_count": s["detalhe_status_count"],
        "qualidade": {
            "registros_nd": s["registros_nd"],
            "percentual_acima_100": s["perc_acima_100"],
            "percentual_abaixo_80": s["perc_abaixo_80"],
            "duplicidades": len(s["duplicidades"]),
        },
        "faixas": s["faixas"],
        "vinculo": ctx.get("vinculo"),
        "alertas": ctx["alertas"],
        "alertas_nome_total": len(ctx["alertas_nome"]),
        "amostra": ctx["amostra"],
    }


# ---------------------------------------------------------------------------
# Escrita futura (IDEB-03B) — preparada mas NÃO executada nesta etapa
# ---------------------------------------------------------------------------
UPSERT_SQL = """
INSERT INTO ideb_resultados (
    ano, codigo_inep, school_id, nome_escola_origem, etapa,
    total_avaliado, percentual_avaliado, proficiencia_portugues,
    proficiencia_matematica, fluxo_indicador_rendimento, ideb,
    status_ideb, detalhe_status_ideb, status_vinculo,
    fonte_arquivo, fonte_inep_url, import_batch_id, updated_at
) VALUES (
    %(ano)s, %(codigo_inep)s, %(school_id)s, %(nome_escola_origem)s, %(etapa)s,
    %(total_avaliado)s, %(percentual_avaliado)s, %(proficiencia_portugues)s,
    %(proficiencia_matematica)s, %(fluxo_indicador_rendimento)s, %(ideb)s,
    %(status_ideb)s, %(detalhe_status_ideb)s, %(status_vinculo)s,
    %(fonte_arquivo)s, %(fonte_inep_url)s, %(import_batch_id)s, CURRENT_TIMESTAMP
)
ON CONFLICT (ano, codigo_inep, etapa) DO UPDATE SET
    school_id = EXCLUDED.school_id,
    nome_escola_origem = EXCLUDED.nome_escola_origem,
    total_avaliado = EXCLUDED.total_avaliado,
    percentual_avaliado = EXCLUDED.percentual_avaliado,
    proficiencia_portugues = EXCLUDED.proficiencia_portugues,
    proficiencia_matematica = EXCLUDED.proficiencia_matematica,
    fluxo_indicador_rendimento = EXCLUDED.fluxo_indicador_rendimento,
    ideb = EXCLUDED.ideb,
    status_ideb = EXCLUDED.status_ideb,
    detalhe_status_ideb = EXCLUDED.detalhe_status_ideb,
    status_vinculo = EXCLUDED.status_vinculo,
    fonte_arquivo = EXCLUDED.fonte_arquivo,
    fonte_inep_url = EXCLUDED.fonte_inep_url,
    import_batch_id = EXCLUDED.import_batch_id,
    updated_at = CURRENT_TIMESTAMP
""".strip()


def linha_para_params(linha, ctx):
    """Converte uma linha normalizada nos parâmetros nomeados do UPSERT (IDEB-03B)."""
    return {
        "ano": linha["ano"],
        "codigo_inep": linha["codigo_inep"],
        "school_id": linha.get("school_id"),
        "nome_escola_origem": linha["nome_escola_origem"],
        "etapa": linha["etapa"],
        "total_avaliado": linha["total_avaliado"],
        "percentual_avaliado": linha["percentual_avaliado"],
        "proficiencia_portugues": linha["proficiencia_portugues"],
        "proficiencia_matematica": linha["proficiencia_matematica"],
        "fluxo_indicador_rendimento": linha["fluxo_indicador_rendimento"],
        "ideb": linha["ideb"],
        "status_ideb": linha["status_ideb"],
        "detalhe_status_ideb": linha["detalhe_status_ideb"],
        "status_vinculo": linha.get("status_vinculo", "pendente_validacao"),
        "fonte_arquivo": ctx["fonte_arquivo"],
        "fonte_inep_url": FONTE_INEP_URL,
        "import_batch_id": ctx["batch_id"],
    }


def executar_apply(linhas, ctx, dsn):
    """Executa a carga real (UPSERT) — RESERVADO para IDEB-03B.

    Faz INSERT ... ON CONFLICT (ano, codigo_inep, etapa) DO UPDATE em transação.
    Nunca usa TRUNCATE/DELETE. Esta função NÃO é chamada no escopo IDEB-03A.
    """
    if not dsn:
        raise RuntimeError("--apply requer DSN de banco (DATABASE_URL/DB_DSN). Abortado.")
    if not _HAS_PSYCOPG:
        raise RuntimeError("--apply requer psycopg instalado. Abortado.")

    inseridos = 0
    with psycopg.connect(dsn, connect_timeout=10) as conn:  # type: ignore[union-attr]
        with conn.cursor() as cur:
            for linha in linhas:
                cur.execute(UPSERT_SQL, linha_para_params(linha, ctx))
                inseridos += 1
        conn.commit()
    return inseridos


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def parse_args(argv):
    p = argparse.ArgumentParser(
        description="Importador (dry-run) da base IDEB 2023 para ideb_resultados.",
    )
    p.add_argument(
        "--source",
        default=f"_local/ideb/fontes/{FONTE_ARQUIVO_PADRAO}",
        help="Caminho da planilha IDEB 2023 (.xlsx). Insumo LOCAL, não versionado.",
    )
    p.add_argument("--ano", type=int, default=2023, help="Ano de referência (default: 2023).")
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Modo seguro: analisa e gera relatório SEM escrever no banco (padrão).",
    )
    p.add_argument(
        "--apply",
        action="store_true",
        help="Carga real (RESERVADO para IDEB-03B). Exige também --confirm-apply.",
    )
    p.add_argument(
        "--confirm-apply",
        action="store_true",
        help="Confirmação explícita obrigatória para --apply.",
    )
    p.add_argument("--batch-id", default=None, help="Identificador da carga (auditoria).")
    p.add_argument("--dsn", default=None, help="DSN do PostgreSQL (sobrepõe variáveis de ambiente).")
    p.add_argument(
        "--no-db",
        action="store_true",
        help="Ignora qualquer conexão de banco mesmo em dry-run (não simula match).",
    )
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv if argv is not None else sys.argv[1:])

    # --- Segurança operacional: modo padrão seguro ---
    if not args.apply and not args.dry_run:
        print("[INFO] Nenhum modo informado; assumindo --dry-run (modo seguro).")
        args.dry_run = True

    if args.apply:
        # IDEB-03A NÃO executa carga. A trava abaixo impede --apply nesta etapa.
        if not args.confirm_apply:
            print(
                "[ERRO] --apply exige confirmação explícita --confirm-apply.\n"
                "       A carga real é escopo do IDEB-03B e NÃO deve ser executada agora.",
                file=sys.stderr,
            )
            return 2
        print(
            "[ERRO] Carga real (--apply) está reservada para o incremento IDEB-03B.\n"
            "       Esta versão do importador (IDEB-03A) não autoriza escrita no banco.\n"
            "       Abortado por segurança.",
            file=sys.stderr,
        )
        return 2

    # --- Validação da fonte ---
    if not os.path.isfile(args.source):
        print(
            f"[ERRO] Planilha não encontrada: {args.source}\n"
            "       Coloque a base IDEB 2023 em _local/ (não versionado).",
            file=sys.stderr,
        )
        return 2

    fonte_arquivo = os.path.basename(args.source)
    batch_id = args.batch_id or f"ideb_{args.ano}_{datetime.now():%Y%m%d_%H%M%S}"

    print("=" * 70)
    print("IDEB-03A — Importador (dry-run)")
    print("=" * 70)
    print(f"Fonte ............. {args.source}")
    print(f"Ano ............... {args.ano}")
    print(f"Modo .............. dry-run (sem escrita)")
    print(f"Batch-id .......... {batch_id}")

    # --- Leitura e normalização ---
    try:
        _, registros = carregar_planilha(args.source)
    except KeyError as e:
        print(f"[ERRO] {e}", file=sys.stderr)
        return 2

    linhas = [normalizar_registro(r, args.ano) for r in registros]

    alertas = []
    # Etapas não mapeadas viram alerta (não bloqueio).
    n_desconhecidas = sum(1 for l in linhas if l["etapa"] == "desconhecida")
    if n_desconhecidas:
        alertas.append(f"{n_desconhecidas} registro(s) com etapa não reconhecida (verificar coluna Ensino).")
    n_inep_ausente = sum(1 for l in linhas if not l["codigo_inep"])
    if n_inep_ausente:
        alertas.append(f"{n_inep_ausente} registro(s) sem INEP (erro crítico para carga futura).")

    # --- Conexão opcional para simular match ---
    dsn = "" if args.no_db else resolver_dsn(args.dsn)
    dsn_mascarado = mascarar_dsn(dsn) if dsn else ""
    houve_conexao = False
    schools_por_inep = {}
    if dsn and not args.no_db:
        if not _HAS_PSYCOPG:
            alertas.append("DSN presente, mas psycopg não está instalado; match com schools não simulado.")
        else:
            ineps = {l["codigo_inep"] for l in linhas if l["codigo_inep"]}
            try:
                print(f"DB ................ tentando conexão somente leitura: {dsn_mascarado}")
                schools_por_inep = consultar_schools_por_inep(dsn, ineps)
                houve_conexao = True
                print(f"DB ................ schools consultadas: {len(schools_por_inep)} INEPs encontrados")
            except Exception as e:  # pragma: no cover - depende de ambiente
                # Mascara qualquer detalhe sensível no texto da exceção.
                msg = str(e).replace(dsn, dsn_mascarado) if dsn else str(e)
                alertas.append(f"Falha ao consultar schools (match não executado): {msg[:200]}")
                print(f"DB ................ falha de conexão; seguindo sem match. ({msg[:120]})")
    else:
        print("DB ................ sem conexão (match real não executado)")

    alertas_nome = resolver_vinculo(linhas, schools_por_inep, houve_conexao)

    # --- Estatísticas ---
    stats = calcular_estatisticas(linhas)

    vinculo = None
    if houve_conexao:
        vinculo = {
            "match_inep": sum(1 for l in linhas if l.get("status_vinculo") == "match_inep"),
            "sem_match_inep": sum(1 for l in linhas if l.get("status_vinculo") == "sem_match_inep"),
            "ineps_match": len(
                {l["codigo_inep"] for l in linhas if l.get("status_vinculo") == "match_inep"}
            ),
            "ineps_sem_match": len(
                {l["codigo_inep"] for l in linhas if l.get("status_vinculo") == "sem_match_inep"}
            ),
        }

    ctx = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": args.source,
        "fonte_arquivo": fonte_arquivo,
        "batch_id": batch_id,
        "ano": args.ano,
        "houve_conexao": houve_conexao,
        "dsn_mascarado": dsn_mascarado,
        "stats": stats,
        "vinculo": vinculo,
        "alertas": alertas,
        "alertas_nome": alertas_nome,
        "amostra": amostra_registros(linhas, 10),
    }

    # --- Geração dos relatórios locais ---
    os.makedirs(os.path.dirname(RELATORIO_MD), exist_ok=True)
    with open(RELATORIO_MD, "w", encoding="utf-8") as f:
        f.write(gerar_relatorio_md(ctx))
    with open(RELATORIO_JSON, "w", encoding="utf-8") as f:
        json.dump(gerar_relatorio_json(ctx), f, ensure_ascii=False, indent=2)

    # --- Resumo no console ---
    print("-" * 70)
    print(f"Registros lidos ... {stats['total_registros']}")
    print(f"INEPs únicos ...... {stats['ineps_unicos']}")
    print(f"Com IDEB .......... {stats['com_ideb']}")
    print(f"Sem IDEB .......... {stats['sem_ideb']}")
    print(f"status_ideb ....... {stats['status_count']}")
    print(f"detalhe_status .... {stats['detalhe_status_count']}")
    print(f"ND proficiência ... {stats['registros_nd']}")
    print(f"% > 100 ........... {stats['perc_acima_100']}")
    print(f"% < 80 ............ {stats['perc_abaixo_80']}")
    print(f"Duplicidades ...... {len(stats['duplicidades'])}")
    if houve_conexao:
        print(f"Match schools ..... {vinculo}")
    if alertas:
        print(f"Alertas ........... {len(alertas)}")
        for a in alertas:
            print(f"  - {a}")
    print("-" * 70)
    print(f"Relatório MD ...... {RELATORIO_MD}")
    print(f"Relatório JSON .... {RELATORIO_JSON}")
    print("Nenhuma escrita executada no banco (dry-run).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
